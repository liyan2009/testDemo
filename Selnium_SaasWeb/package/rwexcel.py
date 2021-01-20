#! python3
#-*- coding:utf-8-*-

from openpyxl import load_workbook
import os

'''操作excel，操作大批量数据的时候使用openpyxl'''
class RWexcel(object):
    def __init__(self,xls_name,sheet_name):
        proDir = os.path.split(os.path.realpath(__file__))[0]
        curDir = os.path.dirname(proDir)
        caseFile = os.path.join(curDir, "result")
        self.xlsPath = os.path.join(caseFile, xls_name)
        self.wb=load_workbook(self.xlsPath)
        self.sheet=self.wb.active
        self.sheet_name=sheet_name

    '''根据sheet名称来获取sheet'''
    @property
    def get_Sheet(self):
        self.sheet=self.wb[self.sheet_name]
        return self.sheet

    '''获取总行数'''
    @property
    def get_Rows(self):
        rows=self.sheet.max_row
        return  rows

    '''获取总列数'''
    @property
    def get_Columns(self):
        columns=self.sheet.max_column
        return  columns

    '''保存文件'''
    @property
    def save_Excel(self):
        print(self.xlsPath)
        self.wb.save(self.xlsPath)


    '''获取用例ID'''
    @property
    def get_CaseID(self):
        data=[]
        for i in range(2,self.get_Rows+1):
            data.append(self.get_Sheet.cell(row=i,column=1).value)
        return data

    '''获取用例名称'''
    @property
    def get_CaseName(self):
        data=[]
        for i in range(2,self.get_Rows+1):
            data.append(self.get_Sheet.cell(row=i,column=2).value)
        return data

    '''获取发送请求类型'''
    @property
    def get_Method(self):
        data=[]
        for i in range(2,self.get_Rows+1):
            data.append(self.get_Sheet.cell(row=i,column=3).value)
        return data

    '''获取url地址'''
    @property
    def get_Url(self):
        data=[]
        for i in range(2,self.get_Rows+1):
            data.append(self.get_Sheet.cell(row=i,column=4).value)
        return data

    '''获取参数类型地址'''
    @property
    def get_DataType(self):
        data=[]
        for i in range(2,self.get_Rows+1):
            data.append(self.get_Sheet.cell(row=i,column=5).value)
        return data


    '''获取参数值'''
    @property
    def get_Data(self):
        data=[]
        for i in range(2,self.get_Rows+1):
            data.append(self.get_Sheet.cell(row=i,column=6).value)
        return data

    '''获取检查点'''
    @property
    def get_CheckData(self):
        data=[]
        for i in range(2,self.get_Rows+1):
            data.append(self.get_Sheet.cell(row=i,column=7).value)
        return data

    '''修改'''
    def update_resultAndContent(self,i,result,content):
        # for i in range(2,self.get_Rows+1):
        self.get_Sheet.cell(row=i+2,column=10).value=result
        self.get_Sheet.cell(row=i+2,column=11).value=str(content)


# xls_name = "2.xlsx"
# sheet_name = "Sheet1"
# rw=RWexcel(xls_name,sheet_name)
#
# a=rw.get_Columns
# col=rw.get_Rows
# print(a)
# print(col)
# print(rw.get_CaseID)
